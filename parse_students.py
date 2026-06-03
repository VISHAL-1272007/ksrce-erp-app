import openpyxl, json

wb = openpyxl.load_workbook(r'd:\Admin\college\II CSE(IOT).xlsx')

def safe_float(v, default=0.0):
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(',', '').replace(' ', '')
    try: return float(s)
    except: return default

def safe_int(v, default=0):
    f = safe_float(v, None)
    return int(f) if f is not None else default

def safe_str(v, default=''):
    if v is None: return default
    s = str(v).strip()
    return s if s.lower() not in ['none', 'nan', 'null'] else default

def safe_phone(v):
    if v is None: return ''
    if isinstance(v, float): return str(int(v))
    s = str(v).strip().replace(' ', '')
    return s if s.lower() not in ['none', 'nan', 'null', 'nil', '-'] else ''

ws_students = wb['Student details']
ws_marks    = wb['Marksheet details']
ws_fees     = wb['Fees details( 4 sem)']

rows      = list(ws_students.iter_rows(values_only=True))
mark_rows = list(ws_marks.iter_rows(values_only=True))
fee_rows  = list(ws_fees.iter_rows(values_only=True))

# ── Build marks lookup (header is row index 2) ──────────────────────────────
marks_by_reg = {}
for r in mark_rows[3:]:
    if not r[0]: continue
    reg = str(int(safe_float(r[0]))) if isinstance(r[0], float) else str(r[0]).strip()
    cgpa_vals = [safe_float(v) for v in [r[6], r[8], r[10], r[12]] if safe_float(v) > 0]
    avg_cgpa  = round(sum(cgpa_vals) / len(cgpa_vals), 2) if cgpa_vals else 0.0
    dob_raw   = r[2]
    if dob_raw is None:
        dob = ''
    elif hasattr(dob_raw, 'strftime'):
        dob = dob_raw.strftime('%Y-%m-%d')
    else:
        dob = str(dob_raw).strip()[:10]
    marks_by_reg[reg] = {
        'dob':       dob,
        'sem1_cgpa': safe_float(r[6]),
        'sem2_cgpa': safe_float(r[8]),
        'sem3_cgpa': safe_float(r[10]),
        'sem4_cgpa': safe_float(r[12]),
        'avg_cgpa':  avg_cgpa,
        'arrears':   safe_int(r[13]),
    }

# ── Build fees lookup ────────────────────────────────────────────────────────
fees_by_reg = {}
for r in fee_rows[3:]:
    if not r[0]: continue
    reg = str(int(safe_float(r[0]))) if isinstance(r[0], float) else str(r[0]).strip()
    fees_by_reg[reg] = {
        'totalFees':   safe_float(r[3]),
        'paidFees':    safe_float(r[4]),
        'balanceFees': safe_float(r[5]),
    }

# ── Parse student rows (data starts at row index 4) ─────────────────────────
students        = []
users           = []
credentials_list = []

for r in rows[4:]:
    if not r[0]: continue

    reg_raw    = str(int(safe_float(r[0]))) if isinstance(r[0], float) else str(r[0]).strip()
    name_base  = safe_str(r[1]).title()
    initial    = safe_str(r[2])
    full_name  = (name_base + ' ' + initial).strip() if initial else name_base

    phone       = safe_phone(r[3])
    father_name = safe_str(r[4])
    father_ph   = safe_phone(r[5])
    mother_name = safe_str(r[6])
    mother_ph   = safe_phone(r[7])
    father_occ  = safe_str(r[8])
    mother_occ  = safe_str(r[9])
    address     = safe_str(r[10])

    domain_email   = safe_str(r[11]).lower() or f'{reg_raw.lower()}@ksrce.ac.in'
    personal_email = safe_str(r[12]).lower()
    accommodation  = safe_str(r[13])

    bus_no_raw  = safe_str(r[14])
    bus_no      = '' if bus_no_raw.lower() in ['nil', 'none', '-'] else bus_no_raw
    bus_stop_raw = safe_str(r[15])
    bus_stop     = '' if bus_stop_raw.lower() in ['nil', 'none', '-'] else bus_stop_raw
    hostel_room  = str(int(safe_float(r[16]))) if r[16] and isinstance(r[16], float) else safe_str(r[16])
    hostel_room  = '' if hostel_room.lower() in ['nil', 'none', '-'] else hostel_room

    m = marks_by_reg.get(reg_raw, {})
    f = fees_by_reg.get(reg_raw, {})

    student = {
        'studentId':      'STU_' + reg_raw,
        'roll_no':        reg_raw,
        'name':           full_name,
        'email':          domain_email,
        'personalEmail':  personal_email,
        'phone':          phone,
        'departmentId':   'DEPT_CSEIOT',
        'department_id':  'DEPT_CSEIOT',
        'departmentCode': 'CSEIOT',
        'year':           2,
        'section':        'A',
        'cgpa':           m.get('avg_cgpa', 0.0),
        'sem1Cgpa':       m.get('sem1_cgpa', 0.0),
        'sem2Cgpa':       m.get('sem2_cgpa', 0.0),
        'sem3Cgpa':       m.get('sem3_cgpa', 0.0),
        'sem4Cgpa':       m.get('sem4_cgpa', 0.0),
        'arrears':        m.get('arrears', 0),
        'dob':            m.get('dob', ''),
        'fatherName':         father_name,
        'fatherPhone':        father_ph,
        'fatherOccupation':   father_occ,
        'motherName':         mother_name,
        'motherPhone':        mother_ph,
        'motherOccupation':   mother_occ,
        'address':            address,
        'accommodation':      accommodation,
        'busNo':              bus_no,
        'busStop':            bus_stop,
        'hostelRoom':         hostel_room,
        'totalFees':          f.get('totalFees', 0.0),
        'paidFees':           f.get('paidFees', 0.0),
        'balanceFees':        f.get('balanceFees', 0.0),
        'enrolledCourses':    ['CSEIOT_C1', 'CSEIOT_C2', 'CSEIOT_C3'],
        'created_at':         '2024-06-01T00:00:00Z',
        'user_id':            'STU_' + reg_raw,
    }
    students.append(student)

    # Password: First 4 letters of name (no spaces) + last 4 reg digits + @Ksrce
    name_part = name_base.replace(' ', '')[:4] or 'Stud'
    pwd_plain = name_part + reg_raw[-4:] + '@Ksrce'

    user = {
        'id':           'STU_' + reg_raw,
        'role':         'student',
        'portalRole':   'student',
        'name':         full_name,
        'email':        domain_email,
        'phone':        phone,
        'password':     pwd_plain,
        'departmentId': 'DEPT_CSEIOT',
        'studentId':    'STU_' + reg_raw,
    }
    users.append(user)
    credentials_list.append({
        'name':     full_name,
        'roll_no':  reg_raw,
        'email':    domain_email,
        'password': pwd_plain,
    })

print('Total:', len(students), 'students,', len(users), 'users')

# ── Write students.json ──────────────────────────────────────────────────────
with open(r'd:\Admin\college\college\assets\data\students.json', 'w', encoding='utf-8') as fp:
    json.dump(students, fp, indent=2, ensure_ascii=False)
print('students.json written')

# ── Merge with existing non-student users and write users.json ───────────────
existing_users = []
try:
    with open(r'd:\Admin\college\college\assets\data\users.json', 'r', encoding='utf-8') as fp:
        existing_users = json.load(fp)
except:
    pass

admin_users = [u for u in existing_users if u.get('role') not in ['student']]
all_users   = admin_users + users

with open(r'd:\Admin\college\college\assets\data\users.json', 'w', encoding='utf-8') as fp:
    json.dump(all_users, fp, indent=2, ensure_ascii=False)
print('users.json written (' + str(len(all_users)) + ' total)')

# ── Write departments.json ───────────────────────────────────────────────────
departments = [
    {'departmentId':'DEPT_CSE',    'departmentCode':'CSE',    'departmentName':'Computer Science and Engineering',             'level':'UG','established':2001,'intake':180,'nbaStatus':'Accredited'},
    {'departmentId':'DEPT_CSEIOT', 'departmentCode':'CSEIOT', 'departmentName':'CSE - Internet of Things',                    'level':'UG','established':2023,'intake':60, 'nbaStatus':'Not Eligible'},
    {'departmentId':'DEPT_CSECS',  'departmentCode':'CSECS',  'departmentName':'CSE - Cyber Security',                        'level':'UG','established':2024,'intake':60, 'nbaStatus':'Not Eligible'},
    {'departmentId':'DEPT_ECE',    'departmentCode':'ECE',    'departmentName':'Electronics and Communication Engineering',    'level':'UG','established':2001,'intake':150,'nbaStatus':'Accredited'},
    {'departmentId':'DEPT_EEE',    'departmentCode':'EEE',    'departmentName':'Electrical and Electronics Engineering',       'level':'UG','established':2002,'intake':120,'nbaStatus':'Accredited'},
    {'departmentId':'DEPT_IT',     'departmentCode':'IT',     'departmentName':'Information Technology',                      'level':'UG','established':2001,'intake':120,'nbaStatus':'Eligible'},
    {'departmentId':'DEPT_MECH',   'departmentCode':'MECH',   'departmentName':'Mechanical Engineering',                      'level':'UG','established':2005,'intake':120,'nbaStatus':'Accredited'},
    {'departmentId':'DEPT_CE',     'departmentCode':'CE',     'departmentName':'Civil Engineering',                           'level':'UG','established':2002,'intake':60, 'nbaStatus':'Eligible'},
    {'departmentId':'DEPT_MBA',    'departmentCode':'MBA',    'departmentName':'Master of Business Administration',           'level':'PG','established':2006,'intake':60, 'nbaStatus':'Eligible'},
    {'departmentId':'DEPT_MCA',    'departmentCode':'MCA',    'departmentName':'Master of Computer Applications',             'level':'PG','established':2003,'intake':60, 'nbaStatus':'Eligible'},
]

with open(r'd:\Admin\college\college\assets\data\departments.json', 'w', encoding='utf-8') as fp:
    json.dump(departments, fp, indent=2, ensure_ascii=False)
print('departments.json written')

# ── Write faculty.json ───────────────────────────────────────────────────────
faculty = [
    {
        'facultyId':       'FAC_HOD_CSEIOT',
        'name':            'Dr. R. Saravanakumar',
        'email':           'hod.cseiot@ksrce.ac.in',
        'phone':           '9944456200',
        'designation':     'Professor & Head',
        'departmentId':    'DEPT_CSEIOT',
        'isHOD':           True,
        'isClassAdviser':  False,
        'menteeIds':       [],
    },
    {
        'facultyId':       'FAC_CA_CSEIOT',
        'name':            'Ms. K. Priya',
        'email':           'priya.k@ksrce.ac.in',
        'phone':           '9944456201',
        'designation':     'Assistant Professor',
        'departmentId':    'DEPT_CSEIOT',
        'isHOD':           False,
        'isClassAdviser':  True,
        'adviserFor':      {'departmentId':'DEPT_CSEIOT','year':2,'section':'A'},
        'menteeIds':       ['STU_' + s['roll_no'] for s in students[:8]],
    },
    {
        'facultyId':       'FAC_HOD_CSE',
        'name':            'Dr. M. Venkatesan',
        'email':           'hod.cse@ksrce.ac.in',
        'phone':           '9944456202',
        'designation':     'Professor & Head',
        'departmentId':    'DEPT_CSE',
        'isHOD':           True,
        'isClassAdviser':  False,
        'menteeIds':       [],
    },
]

with open(r'd:\Admin\college\college\assets\data\faculty.json', 'w', encoding='utf-8') as fp:
    json.dump(faculty, fp, indent=2, ensure_ascii=False)
print('faculty.json written')

# ── Write classes.json ───────────────────────────────────────────────────────
classes = [
    {
        'classId':        'CLASS_CSEIOT_Y2_A',
        'departmentId':   'DEPT_CSEIOT',
        'departmentCode': 'CSEIOT',
        'year':           2,
        'section':        'A',
        'classAdviserId': 'FAC_CA_CSEIOT',
        'totalStudents':  len(students),
        'classRoomId':    'IOT-LAB-201',
    }
]

with open(r'd:\Admin\college\college\assets\data\classes.json', 'w', encoding='utf-8') as fp:
    json.dump(classes, fp, indent=2, ensure_ascii=False)
print('classes.json written')

# ── Write credentials reference file ────────────────────────────────────────
with open(r'd:\Admin\college\student_credentials.json', 'w', encoding='utf-8') as fp:
    json.dump(credentials_list, fp, indent=2, ensure_ascii=False)
print('student_credentials.json written at d:\\Admin\\college\\')

print()
print('=== Sample Student Login Credentials ===')
for c in credentials_list[:8]:
    print('  ' + c['name'].ljust(20) + ' | ' + c['email'].ljust(45) + ' | ' + c['password'])
